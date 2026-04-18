import numpy as np
import pandas as pd
from openpyxl import load_workbook
import os
import matplotlib.pyplot as plt
import seaborn as sns
import re

from requests import head


class BackTest:
    def __init__(self):
        sheet_names = load_workbook("集思录港股ipo数据.xlsx", data_only=True).sheetnames
        self.集思录ipo数据 = []
        for i in sheet_names[0:2]:
            data = pd.read_excel("集思录港股ipo数据.xlsx", sheet_name=i)
            data["年份"] = i
            self.集思录ipo数据.append(data)
        self.集思录ipo数据 = pd.concat(self.集思录ipo数据)
        self.集思录ipo数据 = self.集思录ipo数据[self.集思录ipo数据["代码"] > 0]
        self.申购中签详情 = pd.concat(
            [pd.read_csv("D:/GitWorkSpace/xianzhi9452/hk-ipo-engine/港股IPO配发结果公告/" + i) for i in
             os.listdir("D:/GitWorkSpace/xianzhi9452/hk-ipo-engine/港股IPO配发结果公告") if i.endswith("csv")] + [
                pd.read_excel("D:/GitWorkSpace/xianzhi9452/hk-ipo-engine/港股IPO配发结果公告/" + i) for i in
                os.listdir("D:/GitWorkSpace/xianzhi9452/hk-ipo-engine/港股IPO配发结果公告") if i.endswith("xls")]
        )
        # self.ipos = []
        # self.errors = []
        # for info in self.集思录ipo数据.itertuples():
        #     申购中签详情_this = self.申购中签详情[self.申购中签详情["代码"] == int(info.代码)]
        #     if len(申购中签详情_this) > 0:
        #         self.ipos.append(pd.concat([info, 申购中签详情_this]))
        #     else:
        #         self.errors.append([info, "miss"])
        # 1. 确保两个表的“代码”列类型一致（都转为整数）
        self.集思录ipo数据["代码"] = self.集思录ipo数据["代码"].astype(int)
        self.申购中签详情["代码"] = self.申购中签详情["代码"].astype(int)

        # 2. 使用左连接 (Left Join)
        # how='left' 表示保留所有“集思录ipo数据”，如果“申购中签详情”没找到对应的，则填 NaN
        self.merged_data = pd.merge(
            self.集思录ipo数据,
            self.申购中签详情,
            on="代码",
            how="left",
            suffixes=('', '_detail')  # 如果有重名列，详情表的列名会带后缀
        )
        self.ipos = self.merged_data[self.merged_data["名称"].notna()]  # 假设“名称”是详情表里的列
        self.ipos = self.ipos[self.ipos['超购倍数'] >= 20].copy()
        self.errors = self.merged_data[self.merged_data["名称"].isna()]

        keep_cols = ["代码", "名称","年份", "申购截止","上市日", "一手资金", "一手","超购倍数", "首日涨幅","申购手数","期望获配手数"]
        self.ipos = self.ipos[keep_cols]


    def printExample(self):
        # --- 设置 Pandas 显示选项，确保对齐和完整显示 ---
        # 显示所有列
        pd.set_option('display.max_columns', None)
        # 确保不换行显示（设置足够大的显示宽度）
        pd.set_option('display.width', 1000)
        # 如果数据中包含中文，为了防止中文对齐失效，需开启以下两个设置
        pd.set_option('display.unicode.ambiguous_as_wide', True)
        pd.set_option('display.unicode.east_asian_width', True)

        try:
            print("正在加载数据...")

            print("\n" + "=" * 50)
            print("【集思录IPO数据示例】")
            print("=" * 50)
            # 打印前5行
            print(self.集思录ipo数据.head())

            print("\n" + "=" * 50)
            print("【申购中签详情示例】")
            print("=" * 50)
            if not self.申购中签详情.empty:
                print(self.申购中签详情.head())
            else:
                print("申购中签详情数据为空，请检查路径。")

            print("\n" + "=" * 50)
            print("【有效IPO】")
            print("=" * 50)
            # 打印前5行
            print(self.ipos)


            # print("\n" + "=" * 50)
            # print("【无效IPO】")
            # print("=" * 50)
            # # 打印前5行
            # print(self.errors)

        except Exception as e:
            print(f"运行出错: {e}")

    def filter_by_capital(self, df, total_capital=200000):
        """
        根据总资金过滤申购档位，并为每只股票保留唯一的最大可申请手数行。
        """
        if df.empty:
            return df

        # 1. 确保数据类型正确
        df = df.copy()  # 避免 SettingWithCopyWarning
        df['一手资金'] = pd.to_numeric(df['一手资金'], errors='coerce')
        df['申购手数'] = pd.to_numeric(df['申购手数'], errors='coerce')

        # 2. 步骤一：过滤掉所有买不起的行
        # 只保留 (一手资金 * 申购手数) <= 20w 的所有可能档位
        mask = (df['一手资金'] * df['申购手数']) <= total_capital
        df_filtered = df[mask].copy()

        if df_filtered.empty:
            print(f"警告：没有任何股票的申购档位在 {total_capital} 预算内。")
            return df_filtered

        # 3. 步骤二：在买得起的档位中，找到每只股票的最大手数
        # 先排序，确保 '申购手数' 是从小到大排列的
        df_filtered = df_filtered.sort_values(by=['代码', '申购手数'], ascending=[True, True])

        # 4. 关键逻辑：按 '代码' 分组，取最后一行 (即该预算下的最大手数)
        # 这样每只股票就只剩下一行数据
        result_df = df_filtered.groupby('代码', as_index=False).last()

        return result_df
    def doTest(self):
        print(123)

    def calculate_expected_profit(self,df):
        # 1. 确保数据类型为数值
        df['期望获配手数'] = pd.to_numeric(df['期望获配手数'], errors='coerce')
        df['一手资金'] = pd.to_numeric(df['一手资金'], errors='coerce')
        df['首日涨幅'] = pd.to_numeric(df['首日涨幅'], errors='coerce')

        # 2. 处理“大于1的当成1处理”
        # .clip(upper=1) 会将所有大于1的值变为1，小于1的保持不变
        adjusted_lots = df['期望获配手数'].clip(upper=1)

        # 3. 计算单只股票收益：期望手数(修正后) * 一手资金 * 首日涨幅
        # 注意：如果数据中的“首日涨幅”是百分数（如10%），需要确认为 0.1 还是 10
        individual_profits = adjusted_lots * df['一手资金'] * df['首日涨幅']

        # 4. 计算综合（总和）
        total_profit = individual_profits.sum()

        return total_profit


    def simulate_random_profit(self, df):
        """
        基于中签概率随机模拟实际收益
        """
        if df.empty:
            return 0

        # 1. 确保数据类型
        df = df.copy()
        df['期望获配手数'] = pd.to_numeric(df['期望获配手数'], errors='coerce').fillna(0)
        df['一手资金'] = pd.to_numeric(df['一手资金'], errors='coerce').fillna(0)
        df['首日涨幅'] = pd.to_numeric(df['首日涨幅'], errors='coerce').fillna(0)

        # 2. 生成随机数 (0.0 到 1.0 之间，数量与 df 行数一致)
        random_thresholds = np.random.rand(len(df))

        # 3. 判定是否中签
        # 概率判定：如果随机数 < 期望获配手数，则中签 (记为 1)，否则为 0
        # 注意：这里隐式处理了 >1 的情况，因为随机数永远 <= 1，
        # 所以如果期望获配是 1.83，(random < 1.83) 永远为 True，即 100% 中签 1 手
        df['是否中签'] = (random_thresholds < df['期望获配手数']).astype(int)

        # 4. 计算随机后的实际收益
        df['随机收益'] = df['是否中签'] * df['一手资金'] * df['首日涨幅']

        # 5. 返回总和
        total_simulated_profit = df['随机收益'].sum()

        return total_simulated_profit, df

    import pandas as pd
    import numpy as np
    def do_simulate_random_profit(self):
        simulated_profits = []
        for i in range(1000):
            profit, _ = self.simulate_random_profit(self.ipos)
            simulated_profits.append(profit)

        print(f"1000次模拟的平均总收益: {np.mean(simulated_profits):.2f}")
        print(f"最差情况收益: {np.min(simulated_profits):.2f}")
        print(f"最好情况收益: {np.max(simulated_profits):.2f}")
        # --- 2. 绘图设置 ---
        plt.rcParams['font.sans-serif'] = ['SimHei']  # 解决中文显示问题
        plt.rcParams['axes.unicode_minus'] = False  # 解决负号显示问题
        plt.figure(figsize=(12, 6))

        # 绘制分布曲线
        sns.histplot(simulated_profits, kde=True, color='skyblue', bins=30, edgecolor='white')

        # 添加均值线和中位数线
        mean_val = np.mean(simulated_profits)
        median_val = np.median(simulated_profits)
        plt.axvline(mean_val, color='red', linestyle='--', label=f'平均收益: {mean_val:.2f}')
        plt.axvline(median_val, color='green', linestyle='-', label=f'中位数收益: {median_val:.2f}')

        # --- 3. 图表装饰 ---
        plt.title('100次港股IPO打新模拟收益分布图', fontsize=15)
        plt.xlabel('累计收益 (HKD)', fontsize=12)
        plt.ylabel('出现频率', fontsize=12)
        plt.legend()
        plt.grid(axis='y', alpha=0.3)

        plt.show()


    def simulate_real_flow(self, default_initial_capital=200000, flag=1):

        # 初始总本金
        initial_capital = default_initial_capital * 10 if flag else default_initial_capital
        df = self.filter_by_capital(self.ipos,initial_capital)
        if df.empty:
            return {"总收益": 0, "最终收益率": 0, "参与项目数": 0}

        df = df.copy()

        # 1. 强力日期清洗
        def parse_ipo_date(date_str, year_str):
            try:
                year = re.search(r'\d{4}', str(year_str)).group(0)
                date_clean = re.search(r'(\d{1,2})-(\d{1,2})', str(date_str))
                if date_clean:
                    return pd.to_datetime(f"{year}-{date_clean.group(1)}-{date_clean.group(2)}")
                return pd.NaT
            except:
                return pd.NaT

        df['申购截止_dt'] = df.apply(lambda x: parse_ipo_date(x['申购截止'], x['年份']), axis=1)
        df['上市日_dt'] = df.apply(lambda x: parse_ipo_date(x['上市日'], x['年份']), axis=1)
        df = df.dropna(subset=['申购截止_dt', '上市日_dt']).sort_values('申购截止_dt')
        # print(df.head())
        locked_funds = []  # 仅记录【中签】导致的资金锁定：[上市日, 锁定金额, 收益]
        total_profit = 0
        executed_count = 0
        break_count = 0
        profit_count = 0
        frozon_date=pd.Timestamp(year=1990, month=1, day=1)
        # 3. 时间轴模拟
        # print(df.shape)
        for _, row in df.iterrows():
            current_date = row['申购截止_dt']
            if frozon_date>current_date:
                break_count+=1
                # print("资金冻结：拒绝打新",row["名称"])
                continue
            # --- B. 判定申购 ---
            # 随机模拟中签
            is_win = np.random.rand() < row['期望获配手数']
            if is_win:
                # 只有中签了，资金才被锁死到上市日
                potential_profit = row['一手资金'] * row['首日涨幅']
                frozon_date = row['上市日_dt']
                locked_funds.append([row['上市日_dt'], potential_profit])
                profit_count+=1
                # print("中签：收益", row['一手资金'] * row['首日涨幅'])
            else:
                # 不中签：资金不占用（或认为瞬时返还），不影响余额，无收益
                # print("未中签：收益", row['一手资金'] * row['首日涨幅'])
                pass
            executed_count += 1

        # --- C. 最后结算 ---
        for _, profit in locked_funds:
            total_profit += profit
        total_profit -= executed_count*100
        final_roi = round(total_profit / default_initial_capital, 4)
        # print({
        #     "总收益": round(total_profit, 2),
        #     "最终收益率": final_roi,
        #     "参与项目数": executed_count,
        #     "中签项目数": profit_count,
        #     "无法申购项目数":break_count
        # })
        return {
            "总收益": round(total_profit, 2),
            "最终收益率": final_roi,
            "参与项目数": executed_count,
            "中签项目数": profit_count,
            "无法申购项目数":break_count
        }
    def do_simulate_real_flow(self, simulation_count = 1000):
        # --- 1. 执行模拟 ---
        results = []

        for i in range(simulation_count):
            res = self.simulate_real_flow(default_initial_capital=200000)
            results.append(res['总收益'])
            if (i + 1) % 10 == 0:
                print(f"已完成 {i + 1} 次模拟...")

        # --- 2. 计算统计指标 ---
        results_array = np.array(results)
        mean_val = np.mean(results_array)
        median_val = np.median(results_array)
        variance_val = np.var(results_array)  # 方差
        std_dev_val = np.std(results_array)  # 标准差

        # --- 3. 绘图逻辑 ---
        plt.rcParams['font.sans-serif'] = ['SimHei']
        plt.rcParams['axes.unicode_minus'] = False

        plt.figure(figsize=(12, 7))

        # 绘制收益波动折线图
        plt.plot(range(1, simulation_count + 1), results, marker='o',
                 linestyle='-', color='#2c7fb8', alpha=0.6, label='单次模拟收益')

        # 绘制均值线
        plt.axhline(mean_val, color='red', linestyle='--', linewidth=2,
                    label=f'均值 (Mean): {mean_val:.2f}')

        # 绘制中位数线
        plt.axhline(median_val, color='green', linestyle='-', linewidth=2,
                    label=f'中位数 (Median): {median_val:.2f}')

        # 绘制标准差区间 (均值 ± 1个标准差)
        # 统计学上，约 68% 的模拟结果会落在红色阴影区域内
        plt.fill_between(range(1, simulation_count + 1),
                         mean_val - std_dev_val, mean_val + std_dev_val,
                         color='red', alpha=0.1, label=f'标准差区间 (±1 Std Dev)')

        # 图表装饰
        plt.title(f"打新策略蒙特卡洛模拟 - 风险与收益分析 ({simulation_count}次)", fontsize=15)
        plt.xlabel("模拟次数 (Iteration)", fontsize=12)
        plt.ylabel("总收益 (HKD)", fontsize=12)
        plt.grid(True, linestyle=':', alpha=0.5)

        # 在图表右上角打印统计信息框
        stats_text = (f'均值: {mean_val:.2f}\n'
                      f'中位数: {median_val:.2f}\n'
                      f'方差: {variance_val:.2f}\n'
                      f'标准差: {std_dev_val:.2f}\n'
                      f'变异系数: {std_dev_val / mean_val:.2%}')
        plt.text(simulation_count * 1.02, mean_val, stats_text,
                 bbox=dict(facecolor='white', alpha=0.5), verticalalignment='center')

        plt.legend(loc='upper left', bbox_to_anchor=(0, 1.15), ncol=4)
        plt.tight_layout()
        plt.show()

        # 控制台详细输出
        print(f"\n" + "=" * 40)
        print(f"{'统计指标':<15} | {'数值':<15}")
        print("-" * 40)
        print(f"{'平均总收益':<15} | {mean_val:>15.2f}")
        print(f"{'中位数收益':<15} | {median_val:>15.2f}")
        print(f"{'收益方差':<15} | {variance_val:>15.2f}")
        print(f"{'标准差 (波动)':<15} | {std_dev_val:>15.2f}")
        print(f"{'变异系数 (CV)':<15} | {std_dev_val / mean_val:>15.2%}")
        print("=" * 40)
    def do_simulate_real_flow_v2(self, simulation_count = 1000):
        # --- 1. 执行模拟 ---
        results = []

        for i in range(simulation_count):
            # 直接获取函数返回的字典
            res = self.simulate_real_flow(default_initial_capital=200000)

            # 提取返回结果中的“最终收益率”，并转换为百分比格式
            # 假设返回的 0.05 代表 5%
            roi_percentage = res['最终收益率'] * 100
            results.append(roi_percentage)

            if (i + 1) % 10 == 0:
                print(f"已完成 {i + 1} 次模拟...")

        # --- 2. 统计学计算 ---
        results_array = np.array(results)
        mean_val = np.mean(results_array)  # 均值
        median_val = np.median(results_array)  # 中位数
        variance_val = np.var(results_array)  # 方差
        std_dev_val = np.std(results_array)  # 标准差

        # --- 3. 绘图逻辑 ---
        plt.rcParams['font.sans-serif'] = ['SimHei']
        plt.rcParams['axes.unicode_minus'] = False

        plt.figure(figsize=(12, 7))

        # 绘制收益率走势
        plt.plot(range(1, simulation_count + 1), results, marker='o',
                 linestyle='-', color='#007acc', alpha=0.7, label='单次模拟收益率')

        # 绘制均值和中位数参考线
        plt.axhline(mean_val, color='red', linestyle='--', linewidth=2,
                    label=f'均值 (Mean): {mean_val:.2f}%')
        plt.axhline(median_val, color='green', linestyle='-', linewidth=2,
                    label=f'中位数 (Median): {median_val:.2f}%')

        # 绘制标准差区间 (展示波动范围)
        plt.fill_between(range(1, simulation_count + 1),
                         mean_val - std_dev_val, mean_val + std_dev_val,
                         color='red', alpha=0.1, label='标准差区间 (±1 Std Dev)')

        # 图表装饰
        plt.title(f"港股打新策略 - {simulation_count}次模拟收益率(ROI)走势", fontsize=15)
        plt.xlabel("模拟次数", fontsize=12)
        plt.ylabel("收益率 (%)", fontsize=12)
        plt.grid(True, linestyle=':', alpha=0.5)

        # 在侧边显示统计看板
        stats_box = (f'【统计汇总】\n'
                     f'平均收益率: {mean_val:.2f}%\n'
                     f'中位数: {median_val:.2f}%\n'
                     f'收益方差: {variance_val:.4f}\n'
                     f'收益标准差: {std_dev_val:.2f}%\n'
                     f'变异系数(CV): {std_dev_val / mean_val:.2f}')
        plt.text(simulation_count * 1.02, mean_val, stats_box,
                 bbox=dict(facecolor='white', alpha=0.8, edgecolor='gray'),
                 verticalalignment='center')

        plt.legend(loc='upper left', bbox_to_anchor=(0, 1.12), ncol=4)
        plt.tight_layout()
        plt.show()

        # --- 4. 控制台详细报告 ---
        print("\n" + "=" * 45)
        print(f"{'收益率分析指标':<20} | {'数值':<15}")
        print("-" * 45)
        print(f"{'平均收益率':<20} | {mean_val:>14.2f}%")
        print(f"{'中位数收益率':<20} | {median_val:>14.2f}%")
        print(f"{'收益率方差':<20} | {variance_val:>15.4f}")
        print(f"{'收益率标准差':<20} | {std_dev_val:>14.2f}%")
        print(f"{'变异系数 (CV)':<20} | {std_dev_val / mean_val:>15.2f}")
        print("=" * 45)

    def do_full_analysis(self, simulation_count=1000, default_capital=200000):
        """
        一键生成金额和收益率双重分析图
        """
        # --- 1. 执行模拟并收集数据 ---
        amounts = []
        rois = []

        print(f"开始执行 {simulation_count} 次深度模拟...")
        for i in range(simulation_count):
            res = self.simulate_real_flow(default_initial_capital=default_capital)
            amounts.append(res['总收益'])
            rois.append(res['最终收益率'] * 100)  # 转化为百分比

            if (i + 1) % 100 == 0:
                print(f"进度: {i + 1}/{simulation_count}")

        # --- 2. 准备绘图数据 ---
        data_groups = [
            {'data': np.array(amounts), 'title': '总收益 (Amount)', 'unit': 'HKD', 'color': '#2c7fb8'},
            {'data': np.array(rois), 'title': '收益率 (ROI)', 'unit': '%', 'color': '#008080'}
        ]

        # --- 3. 绘图逻辑：创建 2 行 1 列的子图 ---
        plt.rcParams['font.sans-serif'] = ['SimHei']
        plt.rcParams['axes.unicode_minus'] = False
        fig, axes = plt.subplots(2, 1, figsize=(12, 12))

        for idx, group in enumerate(data_groups):
            ax = axes[idx]
            data = group['data']
            mean_val = np.mean(data)
            median_val = np.median(data)
            std_val = np.std(data)

            # 绘制走势线
            ax.plot(range(1, simulation_count + 1), data, marker='o', markersize=2,
                    linestyle='-', color=group['color'], alpha=0.5, label='模拟结果')

            # 均值线 & 中位数线
            ax.axhline(mean_val, color='red', linestyle='--', linewidth=1.5,
                       label=f'均值: {mean_val:.2f}{group["unit"]}')
            ax.axhline(median_val, color='green', linestyle='-', linewidth=1.5,
                       label=f'中位数: {median_val:.2f}{group["unit"]}')

            # 标准差阴影
            ax.fill_between(range(1, simulation_count + 1), mean_val - std_val, mean_val + std_val,
                            color='red', alpha=0.1, label='±1标准差区间')

            # 图表装饰
            ax.set_title(f"打新策略 - {group['title']} 走势分析", fontsize=14)
            ax.set_ylabel(f"{group['title']} ({group['unit']})")
            ax.grid(True, linestyle=':', alpha=0.5)
            ax.legend(loc='upper left', ncol=3, fontsize=9)

            # 统计看板（放在每张图的右上角）
            cv = (std_val / mean_val) if mean_val != 0 else 0
            stats_text = (f'均值: {mean_val:.2f}{group["unit"]}\n'
                          f'中位数: {median_val:.2f}{group["unit"]}\n'
                          f'标准差: {std_val:.2f}{group["unit"]}\n'
                          f'变异系数: {cv:.2%}')
            ax.text(1.01, 0.5, stats_text, transform=ax.transAxes,
                    bbox=dict(facecolor='white', alpha=0.8), verticalalignment='center')

        plt.xlabel("模拟次数 (Iteration)", fontsize=12)
        plt.tight_layout()
        plt.show()

        # --- 4. 控制台详细汇总报告 ---
        # 计算统计量
        stats_data = {
            "指标": ["均值 (Mean)", "中位数 (Median)", "方差 (Variance)", "标准差 (Std Dev)", "变异系数 (CV)"],
            "总收益 (HKD)": [
                f"{np.mean(amounts):.2f}",
                f"{np.median(amounts):.2f}",
                f"{np.var(amounts):.2f}",
                f"{np.std(amounts):.2f}",
                f"{np.std(amounts) / np.mean(amounts):.2%}"
            ],
            "收益率 (%)": [
                f"{np.mean(rois):.2f}%",
                f"{np.median(rois):.2f}%",
                f"{np.var(rois):.4f}",
                f"{np.std(rois):.2f}%",
                f"{np.std(rois) / np.mean(rois):.2%}"
            ]
        }

        print("\n" + "=" * 60)
        print(f"{'打新模拟全量统计报告 (' + str(simulation_count) + '次)':^60}")
        print("-" * 60)
        print(f"{'统计项':<15} | {'总收益 (HKD)':<18} | {'收益率 (%)':<15}")
        print("-" * 60)
        for i in range(len(stats_data["指标"])):
            print(
                f"{stats_data['指标'][i]:<12} | {stats_data['总收益 (HKD)'][i]:>18} | {stats_data['收益率 (%)'][i]:>15}")
        print("=" * 60)
if __name__ == "__main__":
    self = BackTest()

    # self.printExample()

    # print(self.ipos[self.ipos['代码'] == 1768])

    expected_sum = self.calculate_expected_profit(self.ipos)
    print(f"策略综合预期收益总和: {expected_sum:.2f}")
    self.do_full_analysis(1000,100000)
