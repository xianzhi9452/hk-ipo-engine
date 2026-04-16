import numpy as np
import pandas as pd
from openpyxl import load_workbook
import os
class hk_ipo:
    def __init__(self):
        self.超购级别收益率 = [
            ["5000 - 10000", 0.64, 1.08, 1.41, 1.79, 2.04, 2.32, 2.68, 2.94, 3.17, 3.41, 3.65, 5.87],
            ["1800 - 4000 ", 1.38, 2.30, 2.89, 3.30, 3.61, 4.12, 4.48, 4.76, 5.07, 5.33, 6.42, 9.3],
            ["700 - 2000  ", 2.27, 2.97, 3.61, 3.99, 4.43, 4.78, 5.26, 5.65, 6.07, 6.43, 6.84, 10.2],
            ["200 - 700   ", 3.46, 4.25, 4.87, 5.36, 5.86, 6.43, 6.88, 7.37, 7.86, 8.37, 8.74, 11.36],
            ["50 - 200    ", 0.09, -0.48, -0.19, -0.22, 1.36, 2.28, 2.44, 3.61, 3.88, 3.92, 5.82, 8.11],
        ]

        self.收益率历史参考 = pd.read_excel("详细申购手数收益.xlsx")

    def 港股打新交易(self):
        ai_promot = '''
        代码：股票代码
        名称	：股票名称
        一手股数：
        申购手数：
        申购股数：
        分配基准描述
        期望获配股数：基础获配 + 额外获配概率 × 额外股数
        期望获配手数：期望获配股数/一手股数
        以上为表头 和 对应一些描述
        注意：24,885名申请人中的1,245名获得200股股份 的描述 意味着 期望获配股数=10.006027727546716（计算逻辑为200*1,245/24,885）
        请将pdf 的中签明细 整理成csv文件 命名规则为 pdf名称+"_"+股票代码+"_"+股票名称
        直接给出csv 文件下载链接就行 不用说别的
        '''
        s_config = [
            ["5000 - 10000", 0.647, 1.08, 1.41, 1.79],
            ["1800 - 4000", 1.38, 2.3, 2.89, 3.3],
            ["700 - 2000", 2.27, 2.97, 3.61, 3.99],
            ["200 - 700", 3.46, 4.25, 4.87, 5.36],
            ["50 - 500", 0.65, 1.61, 3.16, 4.68],
        ]

        sheet_names = load_workbook("集思录港股ipo数据.xlsx", data_only=True).sheetnames
        集思录ipo数据 = []
        for i in sheet_names[:2]:
            data = pd.read_excel("集思录港股ipo数据.xlsx", sheet_name=i)
            data["年份"] = i
            集思录ipo数据.append(data)
        集思录ipo数据 = pd.concat(集思录ipo数据)
        集思录ipo数据 = 集思录ipo数据[集思录ipo数据["代码"] > 0]
        申购中签详情 = pd.concat(
            [pd.read_csv("D:/GitWorkSpace/xianzhi9452/hk-ipo-engine/港股IPO配发结果公告/" + i) for i in os.listdir("D:/GitWorkSpace/xianzhi9452/hk-ipo-engine/港股IPO配发结果公告") if i.endswith("csv")] + [
                pd.read_excel("D:/GitWorkSpace/xianzhi9452/hk-ipo-engine/港股IPO配发结果公告/" + i) for i in os.listdir("D:/GitWorkSpace/xianzhi9452/hk-ipo-engine/港股IPO配发结果公告") if i.endswith("xls")]
        )
        issues = []
        profits = []
        hands = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100, 200]
        for info in 集思录ipo数据.itertuples():
            申购中签详情_this = 申购中签详情[申购中签详情["代码"] == int(info.代码)]
            if len(申购中签详情_this) > 0:
                one_hit_mum_ = 申购中签详情_this[申购中签详情_this['申购手数'] == 1].iloc[0].期望获配手数
                one_hit_mum = info.一手
                if not (0.95 < one_hit_mum_ / one_hit_mum < 1.05 or abs(one_hit_mum_ - one_hit_mum) < 0.0001):
                    print(one_hit_mum_, one_hit_mum, [int(info.代码), info.名称])
                    issues.append([int(info.代码), info.名称, "error", one_hit_mum_, one_hit_mum])
                elif len(set(申购中签详情_this['申购手数'])) != len(申购中签详情_this['申购手数']):
                    issues.append([int(info.代码), info.名称, "申购手数error", one_hit_mum_, one_hit_mum])
                else:
                    申购中签详情_this = {i.申购手数: i.期望获配手数 for i in 申购中签详情_this.itertuples()}
                    base_info = [info.代码, info.名称, info.申购起始, info.一手资金, info.承销商, info.首日涨幅, info.超购倍数, info.AH,
                                 info.发行机制,
                                 info.配售占比, info.回拨至, info.国际配售超购倍数]
                    all_values_1 = [申购中签详情_this.get(i, np.nan) * info.一手资金 * info.首日涨幅 / info.一手资金 for i in hands]
                    all_values_2 = [申购中签详情_this.get(i, np.nan) for i in hands]
                    s_1 = pd.Series(all_values_1)
                    s_interpolated_1 = list(s_1.interpolate(method='linear'))
                    s_2 = pd.Series(all_values_2)
                    s_interpolated_2 = list(s_2.interpolate(method='linear'))
                    profits.append(base_info + s_interpolated_1 + s_interpolated_2)
            else:
                issues.append([info, "miss"])
        for i in issues:
            print(i)
        profits = pd.DataFrame(
            profits,
            columns=["代码", "名称", "申购起始", "一手资金", "承销商", "首日涨幅", "超购倍数", "AH", "发行机制", "配售占比", "回拨至", "国际配售超购倍数"]
                    + ["{}_gain".format(i) for i in hands]
                    + ["{}_prob".format(i) for i in hands])
        profits.to_excel("详细申购手数收益.xlsx")

    def 收益率预测(self, target):
        # AH = target["AH"]
        超购倍数 = target["超购倍数"]
        实际超购倍数近似 = abs(
            np.log(np.array(self.收益率历史参考['超购倍数'] / (self.收益率历史参考['回拨至'] / self.收益率历史参考['配售占比']) / 超购倍数)))
        sim_select = []
        init_bar = 0.2
        while len(sim_select) < 0.05 * len(self.收益率历史参考):
            sim_select = np.where(实际超购倍数近似 < init_bar)[0]
            init_bar += 0.05
        consider_data = self.收益率历史参考.iloc[sim_select]
        hands_gain = consider_data[
            ['1_gain', '2_gain', '3_gain', '4_gain', '5_gain', '6_gain', '7_gain', '8_gain', '9_gain', '10_gain',
             '20_gain', '30_gain', '40_gain', '50_gain', '60_gain', '70_gain', '80_gain', '90_gain', '100_gain',
             '200_gain']].mean()
        print(hands_gain)
        if False:
            aaa = ['1_gain', '2_gain', '3_gain', '4_gain', '5_gain', '6_gain', '7_gain', '8_gain', '9_gain',
                   '10_gain',
                   '20_gain', '30_gain', '40_gain', '50_gain',
                   '60_gain',
                   '70_gain', '80_gain', '90_gain', '100_gain',
                   '200_gain']
            港股打新连续超购倍数与收入收益率参考 = []
            超购倍数 = 20
            while 超购倍数 < 20000:
                实际超购倍数近似 = abs(
                    np.log(np.array(self.收益率历史参考['超购倍数'] / (self.收益率历史参考['回拨至'] / self.收益率历史参考['配售占比']) / 超购倍数)))
                名义超购倍数近似 = abs(np.log(np.array(self.收益率历史参考['超购倍数'] / 超购倍数)))
                sim_select = []
                init_bar = 0.2
                while len(sim_select) < 0.05 * len(self.收益率历史参考):
                    sim_select = np.where(np.logical_and(实际超购倍数近似 < init_bar, 名义超购倍数近似 < init_bar))[0]
                    init_bar += 0.05
                consider_data = self.收益率历史参考.iloc[sim_select]
                hands_gain = consider_data[
                    ['1_gain', '2_gain', '3_gain', '4_gain', '5_gain', '6_gain', '7_gain', '8_gain', '9_gain',
                     '10_gain',
                     '20_gain', '30_gain', '40_gain', '50_gain', '60_gain', '70_gain', '80_gain', '90_gain', '100_gain',
                     '200_gain']].mean()
                hands_risk = consider_data[
                    ['1_gain', '2_gain', '3_gain', '4_gain', '5_gain', '6_gain', '7_gain', '8_gain', '9_gain',
                     '10_gain',
                     '20_gain', '30_gain', '40_gain', '50_gain', '60_gain', '70_gain', '80_gain', '90_gain', '100_gain',
                     '200_gain']].std()
                港股打新连续超购倍数与收入收益率参考.append([超购倍数, len(sim_select)] + list(hands_gain) + list(hands_risk))
                超购倍数 *= 1.05
            港股打新连续超购倍数与收入收益率参考 = pd.DataFrame(港股打新连续超购倍数与收入收益率参考,
                                              columns=["超购倍数", "样本数量"] + aaa + [i.replace("gain", "risk") for i in aaa])
            港股打新连续超购倍数与收入收益率参考.to_excel("超购倍数与风险收益参考.xlsx")

    def find_maximal_combinations(self, hold_cash, targets):
        """
        递归找出所有最大购买方案（买完后剩余现金无法再买任何一个标的的任何一个单位）

        参数:
        hold_cash: 持有现金总额
        targets: 各标的的字典，包含单价和超购级别等信息

        返回:
        所有满足条件的购买方案，每个方案包含各标的的购买数量
        """
        # 提取标的名称和单价
        items = list(targets.keys())
        prices = [targets[item]["单价"] for item in items]
        n = len(items)
        results = []

        def dfs(start, quantities, remaining_cash):
            """
            深度优先递归搜索
            参数:
            start: 从哪个标的开始考虑
            quantities: 当前各标的的购买数量列表
            remaining_cash: 剩余现金
            """
            # 检查是否无法再添加任何标的的任何一个
            can_buy_any = False
            for i in range(n):
                if prices[i] <= remaining_cash:
                    can_buy_any = True
                    # 如果从当前标的开始，可以继续购买
                    if i >= start:
                        new_quantities = quantities.copy()
                        new_quantities[i] += 1
                        dfs(i, new_quantities, remaining_cash - prices[i])

            # 如果无法购买任何标的，且总花费 > 0，则记录当前方案
            if not can_buy_any and sum(quantities) > 0:
                results.append(quantities.copy())

        # 初始化数量数组
        initial_quantities = [0] * n
        dfs(0, initial_quantities, hold_cash)

        # 格式化输出
        formatted = []
        for q in results:
            total_cost = sum(q[i] * prices[i] for i in range(n))
            remaining = hold_cash - total_cost
            formatted.append({
                'quantities': {items[i]: q[i] for i in range(n) if q[i] > 0},  # 只显示购买数量>0的标的
                'quantities_list': q,
                'total_cost': total_cost,
                'remaining_cash': remaining,
                'total_count': sum(q)
            })

        # 按总购买数量降序排序
        formatted.sort(key=lambda x: x['total_count'], reverse=True)
        # 去重（相同数量组合可能重复）
        unique_formatted = []
        seen = set()
        for f in formatted:
            key = tuple(f['quantities_list'])
            if key not in seen:
                seen.add(key)
                unique_formatted.append(f)

        return unique_formatted

    def 现金最佳策略(self, hold_cash, targets):
        combinations = self.find_maximal_combinations(hold_cash, targets)
        result = []
        for combination in combinations:
            total_income = 0
            rrrr = {'quantities': []}
            for t, num in combination['quantities'].items():
                gain_level = self.超购级别收益率[targets[t]["超购级别"]]
                level = num // 10
                if 0 < level < 10:
                    gain = (gain_level[2 + level] - gain_level[1 + level]) / 10 * (num - 10 * level) + gain_level[
                        1 + level]
                elif level == 0:
                    gain = (gain_level[2] - gain_level[1]) / 10 * (num - 1) + gain_level[1]
                else:
                    gain = (gain_level[12] - gain_level[11]) / 100 * (num - 100) + gain_level[11]
                gain = gain / num
                income = num * targets[t]["单价"] * gain / 100
                rrr = {
                    "name": t,
                    "单价": targets[t]["单价"],
                    "申购数量": num,
                    "总价": num * targets[t]["单价"],
                    "收益率": gain,
                    "收益": income
                }
                rrrr['quantities'].append(rrr)
                total_income += income
            rrrr["total_income"] = total_income
            result.append([total_income, rrrr])
        result.sort(reverse=True)
        result = [i[1] for i in result]
        # for i in result:
        #     print(i['total_income'])
        #     for ii in i['quantities']:
        #         print(ii)
        return result

    def 致富融资最佳策略(self, hold_cash, targets):
        combinations = self.find_maximal_combinations(hold_cash, targets)
        result = []
        for combination in combinations:
            total_income = 0
            rrrr = {'quantities': []}
            for t, num_ in combination['quantities'].items():
                gain_level = self.超购级别收益率[targets[t]["超购级别"]]
                if "现金认购":
                    num = num_
                    level = num // 10
                    if 0 < level < 10:
                        gain = (gain_level[2 + level] - gain_level[1 + level]) / 10 * (num - 10 * level) + gain_level[
                            1 + level]
                    elif level == 0:
                        gain = (gain_level[2] - gain_level[1]) / 10 * (num - 1) + gain_level[1]
                    else:
                        gain = (gain_level[12] - gain_level[11]) / 100 * (num - 100) + gain_level[11]

                    gain_cash = gain / num
                    income_cash = num * targets[t]["单价"] * gain_cash / 100
                if "融资认购":
                    num = num_ * 10
                    level = num // 10
                    if 0 < level < 10:
                        gain = (gain_level[2 + level] - gain_level[1 + level]) / 10 * (num - 10 * level) + gain_level[
                            1 + level]
                    elif level == 0:
                        gain = (gain_level[2] - gain_level[1]) / 10 * (num - 1) + gain_level[1]
                    else:
                        gain = (gain_level[12] - gain_level[11]) / 100 * (num - 100) + gain_level[11]
                    gain_margin = gain / num
                    fee = 100
                    income_margin = num * targets[t]["单价"] * gain_margin / 100 - fee
                if income_cash >= income_margin:
                    申购方式 = "现金"
                    income = income_cash
                    gain = gain_cash
                    fee = 0
                else:
                    申购方式 = "融资"
                    income = income_margin
                    gain = gain_margin * 10
                    fee = 100
                rrr = {
                    "name": t,
                    "申购方式 ": 申购方式,
                    "单价": targets[t]["单价"],
                    "申购数量": num_,
                    "fee": fee,
                    "总价": num * targets[t]["单价"],
                    "收益率": gain,
                    "收益": income
                }
                rrrr['quantities'].append(rrr)
                total_income += income
            rrrr["total_income"] = total_income
            result.append([total_income, rrrr])
        result.sort(reverse=True)
        result = [i[1] for i in result]
        # for i in result:
        #     print(i['total_income'])
        #     for ii in i['quantities']:
        #         print(ii)
        return result


# 示例测试
if __name__ == "__main__":
    self = hk_ipo()
    # 账户现金持有量
    hold_cash = 100000
    # 待认购标的
    targets = {
        "MANYCORE TECH": {
            "单价": 3810.00,
            "超购级别": 2,
            "AH": 0
        },
        "思格新能": {
            "单价": 32420,
            "超购级别": 2,
            "AH": 0
        },
        "长光辰芯": {
            "单价": 3988.00,
            "超购级别": 3,
            "AH": 0
        }
    }
    results = self.致富融资最佳策略(hold_cash, targets)
    self.港股打新交易()
    self.收益率预测(targets)
    print(results[0])